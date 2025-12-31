import os
import traceback
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
from tkinter import messagebox
from app.constants import (
    STAGES,
    DEPARTMENTS,
    RECORDS_FOLDER
)
from app.storage import (
    load_stage_department_students,
    load_stage_department_attendance,
    save_stage_department_attendance
)

def export_data():
    try:
        today_date = datetime.now().strftime("%Y-%m-%d")

        for stage in STAGES:
            for department in DEPARTMENTS:

                students = load_stage_department_students(stage, department)
                attendance = load_stage_department_attendance(stage, department)

                if not students:
                    continue

                data = []

                for student in students:
                    student_key = f"{student['name']}|{student['stage']}|{student['department']}"

                    if (
                        student_key in attendance
                        and today_date in attendance[student_key]
                    ):
                        records = attendance[student_key][today_date]
                        attendance_times = [
                            r["time"] for r in records if r["type"] == "حضور"
                        ]
                        departure_times = [
                            r["time"] for r in records if r["type"] == "انصراف"
                        ]

                        attendance_time = (
                            "\n".join(attendance_times)
                            if attendance_times else "غياب"
                        )
                        departure_time = (
                            "\n".join(departure_times)
                            if departure_times else
                            ("غياب" if attendance_times == [] else "لم ينصرف")
                        )
                    else:
                        attendance_time = "غياب"
                        departure_time = "غياب"

                    data.append({
                        "number": len(data) + 1,
                        "name": student["name"],
                        "stage": stage,
                        "department": department,
                        "date": today_date,
                        "attendance": attendance_time,
                        "departure": departure_time,
                        "notes": ""
                    })

                
                data.sort(key=lambda x: x["name"])
                for i, item in enumerate(data, 1):
                    item["number"] = i

                folder_path = os.path.join(RECORDS_FOLDER, stage, department)
                os.makedirs(folder_path, exist_ok=True)

                file_path = os.path.join(folder_path, f"{today_date}.xlsx")

                columns = [
                    "ت", "اسم", "المرحلة", "القسم",
                    "تاريخ", "الحضور", "الانصراف", "ملاحظات"
                ]

                rows = [
                    [
                        item["number"],
                        item["name"],
                        item["stage"],
                        item["department"],
                        item["date"],
                        item["attendance"],
                        item["departure"],
                        item["notes"]
                    ]
                    for item in data
                ]

                df = pd.DataFrame(rows, columns=columns)
                df.to_excel(file_path, index=False, engine="openpyxl")

               
                wb = load_workbook(file_path)
                ws = wb.active
                ws.sheet_view.rightToLeft = True

                header_fill_yellow = PatternFill("solid", start_color="FFFF00")
                header_fill_gray = PatternFill("solid", start_color="D3D3D3")
                header_fill_green = PatternFill("solid", start_color="9BBB59")
                header_fill_blue = PatternFill("solid", start_color="87CEEB")

                header_font = Font(bold=True, size=14)
                data_font = Font(size=12)

                thin = Side(style="thin")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                headers = columns

                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                    if header == "ت":
                        cell.fill = header_fill_yellow
                    elif header in ["اسم", "المرحلة", "القسم", "تاريخ", "ملاحظات"]:
                        cell.fill = header_fill_gray
                    elif header == "الحضور":
                        cell.fill = header_fill_green
                    elif header == "الانصراف":
                        cell.fill = header_fill_blue

                data_fill = PatternFill("solid", start_color="F0F8FF")

                for row in ws.iter_rows(
                    min_row=2,
                    max_row=ws.max_row,
                    min_col=1,
                    max_col=8
                ):
                    for cell in row:
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = data_fill
                        cell.border = border

                ws.column_dimensions["A"].width = 5
                ws.column_dimensions["B"].width = 25
                ws.column_dimensions["C"].width = 15
                ws.column_dimensions["D"].width = 15
                ws.column_dimensions["E"].width = 12
                ws.column_dimensions["F"].width = 12
                ws.column_dimensions["G"].width = 12
                ws.column_dimensions["H"].width = 25

                ws.page_margins = PageMargins(
                    left=0.3, right=0.3,
                    top=0.4, bottom=0.4,
                    header=0.3, footer=0.3
                )

                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0

                wb.save(file_path)

        for stage in STAGES:
            for department in DEPARTMENTS:
                save_stage_department_attendance(stage, department, {})

        messagebox.showinfo("✅ نجح", "تم تصدير البيانات بنجاح")

    except Exception as e:
        messagebox.showerror(
            "❌ خطأ",
            f"فشل في تصدير البيانات:\n{str(e)}"
        )
        print(traceback.format_exc())

